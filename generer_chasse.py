import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Créer un nouveau classeur
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Carte au Trésor"

# Titre en A1:J1
title = "🗺️ CHASSE AU TRÉSOR DES NOMBRES – Trouve le coffre en 10 étapes !"
ws.merge_cells("A1:J1")
ws["A1"] = title
ws["A1"].font = Font(name="Arial", size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

# Redimensionner colonnes A à J
for col in range(1, 11):
    ws.column_dimensions[get_column_letter(col)].width = 15

# Redimensionner lignes 1 à 10
for row in range(1, 11):
    ws.row_dimensions[row].height = 30

# Fond océan pour A2:J10
ocean_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for row in range(2, 11):
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col, value="")
        cell.fill = ocean_fill
        cell.border = thin_border

# Étape 1 : Départ A2 (texte) + B2 (formule résultat)
ws["A2"] = "🚢 Capitaine ! 5 + 3 = ? Descends à la ligne résultat !"
ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
ws["A2"].font = Font(bold=True)
ws["B2"] = "=5+3"  # Résultat : 8
ws["B2"].font = Font(bold=True)
ws["B2"].number_format = "0"

# Étape 2 : A8 (texte) + B8 (formule)
ws["A8"] = "Île aux crabes 🦀 ! Soustrais 4 - 1 = ? Saute à la colonne résultat !"
ws["A8"].alignment = Alignment(wrap_text=True, vertical="center")
ws["A8"].font = Font(bold=True)
ws["B8"] = "=4-1"  # Résultat : 3
ws["B8"].font = Font(bold=True)
ws["B8"].number_format = "0"

# Étape 3 : C8 (texte) + D8 (formule)
ws["C8"] = "Caverne sombre ! 7 + 2 = ? Avance à la case droite autant de fois !"
ws["C8"].alignment = Alignment(wrap_text=True, vertical="center")
ws["C8"].font = Font(bold=True)
ws["D8"] = "=7+2"  # Résultat : 9
ws["D8"].font = Font(bold=True)
ws["D8"].number_format = "0"

# Étape 4 : I8 (texte) + J8 (formule)
ws["I8"] = "Presque là ! 10 - 6 = ? Monte à la ligne résultat !"
ws["I8"].alignment = Alignment(wrap_text=True, vertical="center")
ws["I8"].font = Font(bold=True)
ws["J8"] = "=10-6"  # Résultat : 4
ws["J8"].font = Font(bold=True)
ws["J8"].number_format = "0"

# Trésor I4 (texte) + J4 (somme des sauts)
ws["I4"] = "Trésor ! 🎉 Ajoute ton score pour le total magique."
ws["I4"].alignment = Alignment(wrap_text=True, vertical="center")
ws["I4"].font = Font(bold=True, color="FF0000")  # Rouge
ws["J4"] = (
    "=SOMME(B2;B8;D8;J8)"  # Somme : 24 (note : ; au lieu de , pour compatibilité Calc)
)
ws["J4"].font = Font(bold=True, color="FFD700")  # Or
ws["J4"].number_format = "0"
ws["J4"].alignment = Alignment(horizontal="center")

# Décorations
ws["B3"] = "🌊"
ws["D5"] = "🌊"
ws["F6"] = "🏝️"

# Sauvegarder
wb.save("chasse_au_tresor.xlsx")
print("Fichier créé : chasse_au_tresor.xlsx")  # Message de confirmation
