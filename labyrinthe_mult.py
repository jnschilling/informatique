import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import random

# Créer un nouveau classeur
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Labyrinthe Multiplications"

# Titre en A1:H1
title = "🧩 LABYRINTHE DES MULTIPLICATIONS – Révisons les tables de 1 à 10 !"
ws.merge_cells("A1:H1")
ws["A1"] = title
ws["A1"].font = Font(name="Arial", size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

# Redimensionner colonnes A à H (pour 8x8 grille)
for col in range(1, 9):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Redimensionner lignes 1 à 10
for row in range(1, 11):
    ws.row_dimensions[row].height = 25

# En-têtes : Tables de multiplication (x1 à x7 en B à H)
headers = ["Multiplieur gauche", "x1", "x2", "x3", "x4", "x5", "x6", "x7"]
for col, header in enumerate(headers, start=1):
    ws.cell(row=3, column=col, value=header)
    ws.cell(row=3, column=col).font = Font(bold=True)

# Grille de problèmes : 7 lignes x 7 colonnes (A4:H10)
multipliers = list(range(1, 11))  # 1 à 10 pour gauche
problems = []
lefts = []
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
for row in range(4, 11):  # Lignes 4 à 10
    left_mult = random.choice(multipliers)
    lefts.append(left_mult)
    ws.cell(row=row, column=1, value=left_mult).font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    row_problems = []
    for col in range(2, 9):  # Colonnes B à H (x1 à x7)
        right_mult = col - 1  # col=2 ->1, col=8->7
        problem = f"{left_mult} x {right_mult}"
        cell = ws.cell(row=row, column=col, value=problem)
        actual_answer = left_mult * right_mult
        # Ajouter commentaire avec réponse
        comment = Comment(f"Réponse correcte: {actual_answer}", "Enseignant")
        cell.comment = comment
        row_problems.append(actual_answer)
        cell.border = thin_border
    problems.append(row_problems)

# Ajouter une ligne pour réponses (en bas, ligne 11, colonnes B à H)
ws["A11"] = "Tes réponses (écris ici et vérifie!):"
ws["A11"].font = Font(bold=True)
ws["A11"].border = thin_border
for col in range(2, 9):
    cell = ws.cell(row=11, column=col, value="")
    cell.border = thin_border

# Instructions en A13
instructions = """Instructions: 
1. Pour chaque ligne, calcule les multiplications (ex. 5 x 3 = ?).
2. Écris tes réponses dans la ligne 11 (B11 pour première, etc.).
3. Astuce: Double-clic sur une cellule de problème pour voir l'indice (réponse cachée).
4. Objectif: Complète toutes les lignes sans erreur pour 'sortir' du labyrinthe!
Ajoute une colonne I pour =SI(B11 = {réf}, "✓", "✗") pour auto-correction."""
ws.merge_cells("A13:A17")
ws["A13"] = instructions
ws["A13"].alignment = Alignment(wrap_text=True)

# Ajouter des couleurs de fond pour la grille
grid_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
for row in range(4, 11):
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.fill = grid_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Bordures pour en-têtes et réponses
for row in range(3, 4):  # Ligne 3
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border
for row in range(11, 12):  # Ligne 11
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border

# Sauvegarder
wb.save("labyrinthe_multiplications.xlsx")
print("Fichier créé : labyrinthe_multiplications.xlsx")
